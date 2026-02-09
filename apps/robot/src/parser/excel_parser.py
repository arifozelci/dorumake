"""
KolayRobot Excel Parser
Parses order Excel files from Castrol/suppliers
"""

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Dict, Any
import re

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.utils.logger import logger


@dataclass
class OrderItemData:
    """Single order item (product line)"""
    product_code: str
    product_name: Optional[str] = None
    manufacturer_code: Optional[str] = None
    quantity: int = 0
    unit: str = "ADET"
    unit_price: Optional[Decimal] = None
    total_price: Optional[Decimal] = None
    currency: str = "TRY"
    shipment_date: Optional[datetime] = None
    brand: Optional[str] = None
    manufacturer: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "product_code": self.product_code,
            "product_name": self.product_name,
            "manufacturer_code": self.manufacturer_code,
            "quantity": self.quantity,
            "unit": self.unit,
            "unit_price": float(self.unit_price) if self.unit_price else None,
            "total_price": float(self.total_price) if self.total_price else None,
            "currency": self.currency,
            "shipment_date": self.shipment_date.isoformat() if self.shipment_date else None,
            "brand": self.brand,
            "manufacturer": self.manufacturer,
        }


@dataclass
class OrderData:
    """Parsed order data"""
    # Order info
    order_code: str
    order_date: Optional[datetime] = None
    order_type: Optional[str] = None  # "Stock Order", etc.

    # Customer info
    customer_code: Optional[str] = None
    customer_name: Optional[str] = None
    ship_to_code: Optional[str] = None
    shipping_address: Optional[str] = None

    # Items
    items: List[OrderItemData] = field(default_factory=list)

    # Totals
    total_amount: Optional[Decimal] = None
    currency: str = "TRY"

    # Metadata
    source_file: Optional[str] = None
    parsed_at: datetime = field(default_factory=datetime.utcnow)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "order_code": self.order_code,
            "order_date": self.order_date.isoformat() if self.order_date else None,
            "order_type": self.order_type,
            "customer_code": self.customer_code,
            "customer_name": self.customer_name,
            "ship_to_code": self.ship_to_code,
            "shipping_address": self.shipping_address,
            "items": [item.to_dict() for item in self.items],
            "total_amount": float(self.total_amount) if self.total_amount else None,
            "currency": self.currency,
            "source_file": self.source_file,
            "parsed_at": self.parsed_at.isoformat(),
            "item_count": len(self.items),
            "total_quantity": sum(item.quantity for item in self.items),
        }


class ExcelParser:
    """
    Excel parser for order files

    Supports the "Approved Purchase Order" format from Castrol.
    """

    # Column mappings for the standard format
    # Based on: Approved Purchase Order_20251222032456345.xlsx
    COLUMN_MAPPINGS = {
        "Product Code": ["Product Code", "Ürün Kodu", "Parça No"],
        "Product Name": ["Product Name", "Ürün Adı", "Parça Adı"],
        "Manufacturer Code": ["Product Manufacturer Code", "Üretici Kodu"],
        "Quantity": ["Order Quantity", "Miktar", "Adet", "Sipariş Adedi"],
        "Unit": ["Unit", "Birim"],
        "Unit Price": ["Price Value", "Birim Fiyat", "Fiyat"],
        "Currency": ["Currency", "Para Birimi"],
        "Total Price": ["Total Price Without VAT", "Toplam Tutar", "KDV Hariç Tutar"],
        "Shipment Date": ["Shipment Date", "Sevk Tarihi", "Teslimat Tarihi"],
        "Brand": ["Brand", "Marka"],
        "Manufacturer": ["Manufacturer", "Üretici"],
    }

    # Header row keywords to identify header row
    HEADER_KEYWORDS = ["Product Code", "Ürün Kodu", "Order Quantity", "Miktar"]

    def __init__(self):
        self._column_index_cache = {}

    def _find_header_row(self, ws: Worksheet) -> Optional[int]:
        """Find the row containing column headers"""
        for row_idx in range(1, min(20, ws.max_row + 1)):  # Check first 20 rows
            row_values = [str(cell.value or '').strip() for cell in ws[row_idx]]
            row_text = ' '.join(row_values).lower()

            # Check if this row contains header keywords
            matches = sum(1 for kw in self.HEADER_KEYWORDS if kw.lower() in row_text)
            if matches >= 2:  # At least 2 keywords found
                return row_idx

        return None

    def _find_column_index(
        self,
        ws: Worksheet,
        header_row: int,
        field_name: str
    ) -> Optional[int]:
        """Find column index for a field"""
        possible_names = self.COLUMN_MAPPINGS.get(field_name, [field_name])

        for col_idx in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row=header_row, column=col_idx).value or '').strip()

            for name in possible_names:
                if name.lower() == cell_value.lower():
                    return col_idx

        return None

    def _parse_decimal(self, value: Any) -> Optional[Decimal]:
        """Parse a value to Decimal"""
        if value is None:
            return None

        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            if isinstance(value, str):
                # Remove currency symbols and whitespace
                clean = re.sub(r'[^\d.,\-]', '', value)
                # Handle Turkish number format (1.234,56)
                if ',' in clean and '.' in clean:
                    clean = clean.replace('.', '').replace(',', '.')
                elif ',' in clean:
                    clean = clean.replace(',', '.')
                return Decimal(clean) if clean else None
        except:
            pass

        return None

    def _parse_int(self, value: Any) -> int:
        """Parse a value to int"""
        if value is None:
            return 0

        try:
            if isinstance(value, (int, float)):
                return int(value)
            if isinstance(value, str):
                clean = re.sub(r'[^\d]', '', value)
                return int(clean) if clean else 0
        except:
            pass

        return 0

    def _parse_datetime(self, value: Any) -> Optional[datetime]:
        """Parse a value to datetime"""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value

        try:
            if isinstance(value, str):
                # Try common formats
                formats = [
                    "%Y-%m-%d",
                    "%d/%m/%Y",
                    "%d.%m.%Y",
                    "%m/%d/%Y %H:%M:%S %p",
                    "%Y-%m-%d %H:%M:%S",
                ]
                for fmt in formats:
                    try:
                        return datetime.strptime(value.strip(), fmt)
                    except ValueError:
                        continue
        except:
            pass

        return None

    def _extract_metadata(self, ws: Worksheet, header_row: int) -> Dict[str, Any]:
        """Extract order metadata from rows above header"""
        metadata = {}

        for row_idx in range(1, header_row):
            for col_idx in range(1, min(5, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                next_cell = ws.cell(row=row_idx, column=col_idx + 1)

                cell_value = str(cell.value or '').strip()
                next_value = str(next_cell.value or '').strip() if next_cell.value else ''

                # Look for key-value pairs
                cell_lower = cell_value.lower()

                if 'code' in cell_lower or 'kod' in cell_lower:
                    if 'order' in cell_lower or 'sipariş' in cell_lower:
                        metadata['order_code'] = next_value or cell_value
                    elif 'ship' in cell_lower or 'sevk' in cell_lower:
                        metadata['ship_to_code'] = next_value or cell_value
                    elif 'customer' in cell_lower or 'müşteri' in cell_lower:
                        metadata['customer_code'] = next_value or cell_value
                    elif cell_lower.strip().rstrip(':').strip() == 'code' and next_value:
                        # Plain "Code :" without prefix → customer/company code
                        if 'customer_code' not in metadata:
                            metadata['customer_code'] = next_value

                elif 'name' in cell_lower or 'adı' in cell_lower or 'isim' in cell_lower:
                    # "Name :" or "Company Name :" or "Firma Adı :" - all map to customer_name
                    if next_value and 'customer_name' not in metadata:
                        metadata['customer_name'] = next_value

                elif 'date' in cell_lower or 'tarih' in cell_lower:
                    if 'order' in cell_lower or 'sipariş' in cell_lower:
                        parsed = self._parse_datetime(next_value or cell_value)
                        if parsed:
                            metadata['order_date'] = parsed

                elif 'address' in cell_lower or 'adres' in cell_lower:
                    metadata['shipping_address'] = next_value or cell_value

                elif 'type' in cell_lower or 'tip' in cell_lower:
                    if 'order' in cell_lower or 'sipariş' in cell_lower:
                        metadata['order_type'] = next_value or cell_value

        return metadata

    def parse_file(self, file_path: str) -> Optional[OrderData]:
        """
        Parse an Excel file and extract order data

        Args:
            file_path: Path to Excel file

        Returns:
            OrderData object or None if parsing fails
        """
        logger.info(f"Parsing Excel file: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # Use first sheet (usually "orderInfo")
            ws = wb.active

            # Find header row
            header_row = self._find_header_row(ws)
            if not header_row:
                logger.warning(f"Could not find header row in {file_path}")
                # Try to use row 1 as header
                header_row = 1

            logger.debug(f"Found header row at: {header_row}")

            # Extract metadata
            metadata = self._extract_metadata(ws, header_row)

            # Build column index map
            col_map = {}
            for field in self.COLUMN_MAPPINGS.keys():
                idx = self._find_column_index(ws, header_row, field)
                if idx:
                    col_map[field] = idx

            logger.debug(f"Column mappings: {col_map}")

            # Parse items
            items = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                # Get product code
                product_code_col = col_map.get("Product Code")
                if not product_code_col:
                    continue

                product_code = ws.cell(row=row_idx, column=product_code_col).value
                if not product_code:
                    continue  # Skip empty rows

                product_code = str(product_code).strip()

                # Parse item
                item = OrderItemData(product_code=product_code)

                # Product name
                if "Product Name" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Product Name"]).value
                    item.product_name = str(val).strip() if val else None

                # Manufacturer code
                if "Manufacturer Code" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Manufacturer Code"]).value
                    item.manufacturer_code = str(val).strip() if val else None

                # Quantity
                if "Quantity" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Quantity"]).value
                    item.quantity = self._parse_int(val)

                # Unit
                if "Unit" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Unit"]).value
                    item.unit = str(val).strip() if val else "ADET"

                # Unit price
                if "Unit Price" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Unit Price"]).value
                    item.unit_price = self._parse_decimal(val)

                # Currency
                if "Currency" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Currency"]).value
                    if val:
                        currency = str(val).strip().upper()
                        # Normalize currency codes
                        if currency in ['TRL', 'TL', 'TRY']:
                            item.currency = 'TRY'
                        else:
                            item.currency = currency

                # Total price
                if "Total Price" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Total Price"]).value
                    item.total_price = self._parse_decimal(val)

                # Shipment date
                if "Shipment Date" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Shipment Date"]).value
                    item.shipment_date = self._parse_datetime(val)

                # Brand
                if "Brand" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Brand"]).value
                    item.brand = str(val).strip() if val else None

                # Manufacturer
                if "Manufacturer" in col_map:
                    val = ws.cell(row=row_idx, column=col_map["Manufacturer"]).value
                    item.manufacturer = str(val).strip() if val else None

                # Only add items with quantity > 0
                if item.quantity > 0:
                    items.append(item)

            # Create order data
            order = OrderData(
                order_code=metadata.get('order_code', f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}"),
                order_date=metadata.get('order_date'),
                order_type=metadata.get('order_type'),
                customer_code=metadata.get('customer_code'),
                customer_name=metadata.get('customer_name'),
                ship_to_code=metadata.get('ship_to_code'),
                shipping_address=metadata.get('shipping_address'),
                items=items,
                source_file=file_path,
            )

            # Calculate total
            if items:
                total = sum(
                    item.total_price for item in items
                    if item.total_price is not None
                )
                order.total_amount = total if total else None

                # Get currency from first item
                order.currency = items[0].currency

            logger.info(
                f"Parsed order: {order.order_code} - "
                f"{len(items)} items, "
                f"total qty: {sum(item.quantity for item in items)}"
            )

            return order

        except Exception as e:
            logger.error(f"Error parsing Excel file {file_path}: {e}")
            return None

    def parse_multiple_files(self, file_paths: List[str]) -> List[OrderData]:
        """
        Parse multiple Excel files

        Args:
            file_paths: List of file paths

        Returns:
            List of OrderData objects (excluding failed parses)
        """
        orders = []

        for path in file_paths:
            order = self.parse_file(path)
            if order:
                orders.append(order)

        return orders
